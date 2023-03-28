from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from pathlib import Path
from datetime import date
import openpyxl
import pytest
from constants import global_constants

class Test_OdevClass:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(global_constants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()

    def getData(path,sheet):
        excelFile = openpyxl.load_workbook(path)
        selectedSheet = excelFile[sheet]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data

    path="data/invalid_login.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))
    def test_invalid_login(self,username,password):
        usernameInput = self.waitForElementVisible((By.ID,"user-name"))
        usernameInput.send_keys(username)
        passwordInput = self.waitForElementVisible((By.ID,"password"),10)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"


    path="data/locked_out_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))
    def test_locked_out_login(self,username,password):
        userInput = self.waitForElementVisible((By.ID,"user-name"))
        userInput.send_keys(username)
        passwordinput = self.waitForElementVisible((By.ID,"password"))
        passwordinput.send_keys(password)
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-locked-out-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."


    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("password",getData(path,sheet))
    def test_username_required(self,password):
        passwordinput = self.waitForElementVisible((By.ID,"password"))
        passwordinput.send_keys(password)
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-username-required-login.png")
        assert errorMessage.text == "Epic sadface: Username is required"

    

    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username",getData(path,sheet))
    def test_password_required(self,username):
        userInput = self.waitForElementVisible((By.ID,"user-name"))
        userInput.send_keys(username)
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        errorMessage = self.waitForElementVisible((By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]"))
        self.driver.save_screenshot(f"{self.folderPath}/test-password-required-login.png")
        assert errorMessage.text == "Epic sadface: Password is required"

    
    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))    
    def test_login(self,username,password):
        userInput = self.waitForElementVisible((By.ID,"user-name"))
        userInput.send_keys(username)
        passwordinput = self.waitForElementVisible((By.ID,"password"))
        passwordinput.send_keys(password)
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        productList = WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"inventory_item")))
        self.driver.save_screenshot(f"{self.folderPath}/test-login-{username}-{password}.png")
        assert len(productList) == 6

       
    def test_username_password_required(self):
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-username-password-required-login.png")
        btn = self.waitForElementVisible((By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3/button"))
        btn.click()


    path="data/performance_glitch_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet)) 
    def test_performance_glitch_user(self,username,password):
        userInput = self.waitForElementVisible((By.ID,"user-name"))
        userInput.send_keys(username)
        passwordinput = self.waitForElementVisible((By.ID,"password"))
        passwordinput.send_keys(password)
        button = self.waitForElementVisible((By.ID,"login-button"))
        button.click()
        productList = WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"inventory_item")))
        self.driver.save_screenshot(f"{self.folderPath}/performance-glitch-user-login-{username}-{password}.png")
        assert len(productList) == 6

    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))
    def test_add_to_cart(self,username,password):
        self.test_login(username,password)
        productAddtoCartbtn = self.waitForElementVisible((By.ID,"add-to-cart-sauce-labs-backpack"))
        productAddtoCartbtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-add-to-cart-sauce-labs-backpack.png") 
        shoppingcartlink=self.waitForElementVisible((By.CLASS_NAME,"shopping_cart_link"))
        shoppingcartlink.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-shopping-cart-link.png")
        cartitems= WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"cart_item")))
        assert len(cartitems) == 1
        


    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))
    def test_add_to_cart_3_items(self,username,password):
        self.test_login(username,password)
        productAddtoCartbtn1 = self.waitForElementVisible((By.ID,"add-to-cart-sauce-labs-backpack"))
        productAddtoCartbtn1.click()
        productAddtoCartbtn2 = self.waitForElementVisible((By.ID,"add-to-cart-sauce-labs-bolt-t-shirt"))
        productAddtoCartbtn2.click()
        productAddtoCartbtn3 = self.waitForElementVisible((By.ID,"add-to-cart-sauce-labs-onesie"))
        productAddtoCartbtn3.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-add-to-cart-sauce-labs-backpack.png") 
        shoppingcartlink=self.waitForElementVisible((By.CLASS_NAME,"shopping_cart_link"))
        shoppingcartlink.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-shopping-cart-link.png")
        cartitems= WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"cart_item")))
        assert len(cartitems) == 3

    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))
    def test_remove_from_cart_items(self,username,password):
        self.test_add_to_cart_3_items(username,password)
        productRemovetbtn = self.waitForElementVisible((By.ID,"remove-sauce-labs-bolt-t-shirt"))
        productRemovetbtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-shopping-cart-link_removed.png")
        cartitems= WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"cart_item")))
        assert len(cartitems) == 2

    path="data/standard_user.xlsx"
    sheet="Sheet1"
    @pytest.mark.parametrize("username,password",getData(path,sheet))       
    def test_order(self,username,password):
        self.test_add_to_cart(username,password)
        checkout= self.waitForElementVisible((By.ID,"checkout"))
        checkout.click()
        firstname=self.waitForElementVisible((By.ID,"first-name"))
        firstname.send_keys("Kagan")
        lastname=self.waitForElementVisible((By.ID,"last-name"))
        lastname.send_keys("Efe")
        postalcode=self.waitForElementVisible((By.ID,"postal-code"))
        postalcode.send_keys("06796")
        self.driver.save_screenshot(f"{self.folderPath}/test-order-pre-complete.png")
        continuebtn = self.waitForElementVisible((By.ID,"continue"))
        continuebtn.click()
        finish=self.waitForElementVisible((By.ID,"finish"))
        finish.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-order-finished.png")
        backtoproducts=self.waitForElementVisible((By.ID,"back-to-products"))
        backtoproducts.click()
        productList = WebDriverWait(self.driver,5).until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"inventory_item")))
        assert len(productList) == 6



    def waitForElementVisible(self,locator,timeout=5):
         return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(locator))
    
    
        

        
    